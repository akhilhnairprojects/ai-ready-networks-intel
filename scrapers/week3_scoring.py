# scrapers/week3_scoring.py
# WHAT THIS DOES: reads your account list from accounts.xlsx (committed to the
# repo), scores every account, and writes data/week3/accounts.csv for the Week 3
# page. It runs every week.
#
# 
#
# TWO NUMBERS PER ACCOUNT:
#   * SCORE (0-100)      = readiness = weighted sum of the 5 factors. Sets the Tier.
#   * CONFIDENCE (0-100) = how proven that score is. Rises as real sources arrive
#                          (AI Announce each week; the rest as you verify them).

import os
import csv
import time
import datetime
from urllib.parse import quote_plus

import feedparser
from common import now_iso

# --- paths anchored to the repo, so this works no matter where it is run from ---
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                       # repo root (parent of scrapers/)
XLSX = os.path.join(ROOT, "accounts.xlsx")         # your uploaded account list
OUT  = os.path.join(ROOT, "data", "week3", "accounts.csv")

# Weights (add up to 1.0) — these match your spreadsheet's MODEL WEIGHTS
WEIGHTS = {"ai_hiring": 0.25, "ai_announce": 0.25, "cloud": 0.20, "global": 0.20, "dc": 0.10}
FACTORS = list(WEIGHTS.keys())
HALF_LIFE_DAYS = 90     # a signal older than ~3 months is trusted about half as much

# maps spreadsheet column names -> the model's factor keys
SHEET_COLS = {"AI Hiring": "ai_hiring", "AI Announce": "ai_announce",
              "Cloud": "cloud", "Global": "global", "Data Centre": "dc"}

# --- the model ----------------------------------------------------------------
def decay(age_days):
    return 0.5 ** (age_days / HALF_LIFE_DAYS)

def score_account(factors):
    """Readiness 0-100 = weighted sum of factor values. Drives the Tier."""
    return round(sum(WEIGHTS[n] * factors[n]["val"] for n in WEIGHTS))

def confidence_pct(factors):
    """How trustworthy the score is, 0-100. Low for estimates/stale, high for fresh real sources."""
    return round(100 * sum(WEIGHTS[n] * factors[n]["conf"] * decay(factors[n]["age_days"]) for n in WEIGHTS))

def tier(s):
    return "Tier 1" if s >= 80 else "Tier 2" if s >= 60 else "Tier 3"

def borderline(s):
    return "yes" if (abs(s - 80) <= 3 or abs(s - 60) <= 3) else "no"

# --- the ONE live, self-updating signal: AI Announce --------------------------
def refresh_ai_announce(name, ticker, baseline):
    """Search recent AI news for this company and return
    (new_value, source_url, age_days, confidence). Falls back to the baseline
    estimate (with low confidence) if anything goes wrong."""
    try:
        query = quote_plus(f'"{name}" AI')
        url = f"https://news.google.com/rss/search?q={query}&hl=en-US&gl=US&ceid=US:en"
        parsed = feedparser.parse(url)
        entries = parsed.entries[:30]
        if not entries:
            return baseline, "", 60, 0.3         # no news found -> keep estimate, low confidence

        now = datetime.datetime.now(datetime.timezone.utc)
        intensity, newest_age, top_link = 0.0, 999, entries[0].get("link", url)
        for e in entries:
            if e.get("published_parsed"):
                dt = datetime.datetime(*e.published_parsed[:6], tzinfo=datetime.timezone.utc)
                age = max(0, (now - dt).days)
                if age <= 90:
                    intensity += 0.5 ** (age / 45)   # recent items count more
                    newest_age = min(newest_age, age)
            else:
                intensity += 0.3
        intensity = min(1.0, intensity / 8.0)        # ~8 fresh items = full momentum

        # nudge the estimate by at most +/-15 based on this week's momentum
        nudged = max(0, min(100, round(baseline + (intensity - 0.5) * 30)))
        conf = 0.75 if newest_age <= 90 else 0.4     # fresh corroboration => higher confidence
        return nudged, top_link, (newest_age if newest_age < 999 else 60), conf
    except Exception:
        return baseline, "", 60, 0.3

# --- load the account list from your spreadsheet ------------------------------
def load_from_xlsx():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Account Scoring"] if "Account Scoring" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = lambda name: hdr.index(name)
    accounts = []
    for r in rows[1:]:
        if not r or r[col("Account")] in (None, ""):
            continue
        acct = {
            "account": r[col("Account")],
            "ticker": r[col("Ticker")] if "Ticker" in hdr else "",
            "industry": r[col("Industry")] if "Industry" in hdr else "",
            "factors": {},
        }
        for sheet_name, key in SHEET_COLS.items():
            v = r[col(sheet_name)] if sheet_name in hdr else None
            acct["factors"][key] = {
                "val": float(v) if v is not None else 0.0,
                "conf": 0.3, "age_days": 60, "src": "",   # spreadsheet values start as estimates
            }
        accounts.append(acct)
    return accounts

# small fallback so the page still works if accounts.xlsx hasn't been uploaded yet
FALLBACK = [
    {"account": "Company A (sample)", "ticker": "", "industry": "Technology",
     "factors": {k: {"val": v, "conf": 0.3, "age_days": 60, "src": ""} for k, v in
                 [("ai_hiring", 95), ("ai_announce", 92), ("cloud", 90), ("global", 80), ("dc", 70)]}},
    {"account": "Company B (sample)", "ticker": "", "industry": "Financial Services",
     "factors": {k: {"val": v, "conf": 0.3, "age_days": 60, "src": ""} for k, v in
                 [("ai_hiring", 70), ("ai_announce", 65), ("cloud", 72), ("global", 60), ("dc", 50)]}},
]

def load_accounts():
    if os.path.exists(XLSX):
        try:
            accts = load_from_xlsx()
            print(f"Loaded {len(accts)} accounts from accounts.xlsx")
            return accts
        except Exception as e:
            print(f"WARNING: could not read accounts.xlsx ({e}); using fallback sample.")
    else:
        print("WARNING: accounts.xlsx not found in repo root; using fallback sample. "
              "Upload your spreadsheet as accounts.xlsx to score your real accounts.")
    return FALLBACK

# --- build rows + write CSV ---------------------------------------------------
def build_row(acct):
    f = acct["factors"]
    # refresh the AI-Announce factor from live news this week
    val, src, age, conf = refresh_ai_announce(acct["account"], acct.get("ticker", ""),
                                              f["ai_announce"]["val"])
    f["ai_announce"].update({"val": val, "src": src, "age_days": age, "conf": conf})

    s = score_account(f)
    row = {
        "account": acct["account"], "industry": acct.get("industry", ""),
        "score": s, "tier": tier(s), "confidence": confidence_pct(f),
        "borderline": borderline(s), "last_updated": now_iso()[:10],
    }
    for k in FACTORS:
        row[k + "_score"] = round(f[k]["val"])
        row[k + "_conf"] = f[k]["conf"]
        row[k + "_src"] = f[k]["src"]
    return row

def write_csv(path, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    cols = ["account", "industry", "score", "tier", "confidence", "borderline", "last_updated"]
    for k in FACTORS:
        cols += [k + "_score", k + "_conf", k + "_src"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, quoting=csv.QUOTE_ALL)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"Wrote {len(rows)} accounts to {path}")

if __name__ == "__main__":
    accounts = load_accounts()
    rows = []
    for i, a in enumerate(accounts, 1):
        rows.append(build_row(a))
        time.sleep(0.4)                    # be polite to the news service across many accounts
        if i % 25 == 0:
            print(f"  ...scored {i}/{len(accounts)}")
    rows.sort(key=lambda r: r["score"], reverse=True)
    write_csv(OUT, rows)
    t1 = sum(1 for r in rows if r["tier"] == "Tier 1")
    print(f"Week 3 done: {len(rows)} accounts, {t1} in Tier 1.")
